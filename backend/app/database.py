"""
Database configuration and session management
"""
from sqlalchemy import create_engine, text, inspect
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker

from .config import settings

# Create SQLite engine with check_same_thread=False for FastAPI
SQLALCHEMY_DATABASE_URL = settings.database_url

engine = create_engine(
    SQLALCHEMY_DATABASE_URL,
    connect_args={"check_same_thread": False}  # Needed for SQLite
)

# Session factory
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

# Base class for ORM models
Base = declarative_base()


def get_db():
    """
    Dependency that provides a database session.
    Yields a session and ensures it's closed after use.
    """
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def create_tables():
    """Create all database tables and run lightweight migrations for new columns."""
    Base.metadata.create_all(bind=engine)
    _run_migrations()
    _seed_admin_user()
    _seed_distillation_homework()


def _run_migrations():
    """Add missing columns to existing tables (SQLite doesn't support ALTER TABLE via create_all)."""
    inspector = inspect(engine)

    # Migration: add display_name, method_name, token_encrypted to homework_tokens
    if "homework_tokens" in inspector.get_table_names():
        existing = {col["name"] for col in inspector.get_columns("homework_tokens")}
        with engine.begin() as conn:
            if "display_name" not in existing:
                conn.execute(text("ALTER TABLE homework_tokens ADD COLUMN display_name VARCHAR"))
            if "method_name" not in existing:
                conn.execute(text("ALTER TABLE homework_tokens ADD COLUMN method_name VARCHAR"))
            if "token_encrypted" not in existing:
                conn.execute(text("ALTER TABLE homework_tokens ADD COLUMN token_encrypted TEXT"))
            if "student_uid_encrypted" not in existing:
                conn.execute(text("ALTER TABLE homework_tokens ADD COLUMN student_uid_encrypted TEXT"))

    # Migration: add custom_api_key_encrypted to homework_submissions
    if "homework_submissions" in inspector.get_table_names():
        existing = {col["name"] for col in inspector.get_columns("homework_submissions")}
        with engine.begin() as conn:
            if "custom_api_key_encrypted" not in existing:
                conn.execute(text("ALTER TABLE homework_submissions ADD COLUMN custom_api_key_encrypted TEXT"))


def _seed_admin_user():
    """Create default admin user if it doesn't exist."""
    from .models.user import User
    from .models.user import UserStats
    from .core.security import get_password_hash

    db = SessionLocal()
    try:
        admin_email = "yezhuoyang@cs.ucla.edu"
        existing = db.query(User).filter(User.email == admin_email).first()
        if existing:
            # Ensure admin flag is set
            if not existing.is_admin:
                existing.is_admin = True
                db.commit()
            return

        admin = User(
            email=admin_email,
            username="yezhuoyang",
            password_hash=get_password_hash("yezhuoyang"),
            is_admin=True,
            is_active=True,
        )
        db.add(admin)
        db.flush()

        stats = UserStats(user_id=admin.id)
        db.add(stats)
        db.commit()
        print(f"Seeded admin user: {admin_email}")
    except Exception as e:
        db.rollback()
        print(f"Admin seed skipped: {e}")
    finally:
        db.close()


def _seed_distillation_homework():
    """Create or update the default 'distillation' homework, then seed students from qa_uid.xlsx."""
    import json
    from .models.user import User
    from .models.homework import Homework
    from .services.homework_service import create_homework, generate_tokens_for_homework

    ALLOWED_BACKENDS = [
        "ibm_torino", "ibm_fez", "ibm_kingston", "ibm_marrakesh",
        "ibm_boston", "ibm_pittsburgh", "ibm_miami",
    ]

    db = SessionLocal()
    try:
        existing = db.query(Homework).filter(Homework.problem_id == "distillation").first()
        if existing:
            # Ensure allowed_backends stays up to date
            existing.allowed_backends = json.dumps(ALLOWED_BACKENDS)
            db.commit()
            # Still seed students if xlsx exists
            _seed_students_from_xlsx(db, existing)
            # Seed test student
            _seed_test_student(db, existing)
            return

        # Need admin user as creator
        admin = db.query(User).filter(User.email == "yezhuoyang@cs.ucla.edu").first()
        if not admin:
            print("Distillation homework seed skipped: admin user not found")
            return

        # Use IBMQ token from environment/config
        ibmq_token = settings.ibmq_token
        if not ibmq_token:
            print("Distillation homework seed skipped: IBMQ_TOKEN not set")
            return

        hw = create_homework(
            db=db,
            title="Entanglement Distillation",
            description="Implement an entanglement distillation protocol to improve Bell pair fidelity on real quantum hardware.",
            course="CS 238B",
            ibmq_api_key=ibmq_token,
            ibmq_channel=settings.ibmq_channel,
            ibmq_instance=settings.ibmq_instance,
            allowed_backends=ALLOWED_BACKENDS,
            total_budget_seconds=21600,
            num_students=30,
            max_concurrent_jobs=5,
            problem_id="distillation",
            created_by=admin.id,
        )
        print(f"Seeded distillation homework: {hw.id}")
        _seed_students_from_xlsx(db, hw)
        _seed_test_student(db, hw)
    except Exception as e:
        db.rollback()
        print(f"Distillation homework seed skipped: {e}")
    finally:
        db.close()


def _seed_test_student(db, homework):
    """Seed a test student with UID 000000000."""
    from .services.homework_service import generate_tokens_for_homework

    try:
        results = generate_tokens_for_homework(db, homework, [("000000000", "Test")])
        if results and results[0].get("token"):
            print(f"Test student token: {results[0]['token']}")
    except Exception as e:
        print(f"Test student seed failed: {e}")


def _seed_students_from_xlsx(db, homework):
    """Seed students from qa_uid.xlsx if the file exists."""
    import os
    from .services.homework_service import generate_tokens_for_homework

    # Look for qa_uid.xlsx relative to the project root
    xlsx_candidates = [
        os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "qa_uid.xlsx"),
        os.path.join(os.getcwd(), "qa_uid.xlsx"),
    ]

    xlsx_path = None
    for path in xlsx_candidates:
        if os.path.exists(path):
            xlsx_path = path
            break

    if not xlsx_path:
        return

    try:
        import openpyxl
    except ImportError:
        print("Student seed skipped: openpyxl not installed")
        return

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active

        student_entries = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            name_raw = row[0]
            uid_raw = row[1]
            if not name_raw or not uid_raw:
                continue

            # UID may be int or string; ensure it's a string
            uid = str(uid_raw).strip()

            # Name is "LAST, FIRST" -> convert to "First Last"
            name = str(name_raw).strip()
            if "," in name:
                parts = [p.strip() for p in name.split(",", 1)]
                name = f"{parts[1].title()} {parts[0].title()}"
            else:
                name = name.title()

            student_entries.append((uid, name))

        wb.close()

        if student_entries:
            results = generate_tokens_for_homework(db, homework, student_entries)
            new_count = sum(1 for r in results if r.get("token"))
            print(f"Seeded {new_count} students from {os.path.basename(xlsx_path)}")
    except Exception as e:
        print(f"Student seed from xlsx failed: {e}")
